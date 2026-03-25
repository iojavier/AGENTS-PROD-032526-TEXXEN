import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Agents Productivity Daily Summary", layout="wide")

st.title("📊 Agents Productivity Daily Summary")
st.markdown("Upload multiple XLSX files → Daily Agent Productivity Report (Unique RPC by Debtor ID)")

# ========================= SIDEBAR =========================
with st.sidebar:
    st.header("📁 Data Upload")
    uploaded_files = st.file_uploader(
        "Upload one or more .xlsx files",
        type=["xlsx"],
        accept_multiple_files=True
    )
    st.markdown("---")

if not uploaded_files:
    st.info("👈 Please upload your Excel files using the sidebar uploader.")
    st.stop()

# ========================= LOAD =========================
@st.cache_data(show_spinner=False)
def load_files(files):
    return pd.concat([pd.read_excel(f) for f in files], ignore_index=True)

with st.spinner("Processing files..."):
    df = load_files(uploaded_files)

# ========================= CONSTANTS =========================
EXCLUDED_SUBSTATUSES = {
    "BUSY TONE", "NIS/OOCA", "NO ANSWER", "NO ANSWER_SENT PAYMENT REMINDER",
    "ADC", "NA", "AB", "BUSY_OOCA", "NIS", "BP",
    "LETTER SENT - MANUAL AGENT SMS", "LETTER SENT - MANUAL AGENT EMAIL",
    "KEEPS ON RINGING", "BUSY", "NEGATIVE",
    "MANUAL AGENT EMAIL", "EMAIL", "CALL BARRED"
}

# ========================= PREP =========================
df = df.rename(columns={
    'barcodeDate': 'Date',
    'agent': 'CMS User'
})

df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date
df['CMS User'] = df['CMS User'].fillna('Unknown').astype(str)
df['debtorId'] = df['debtorId'].fillna('').astype(str)

# Normalize text
for col in ['contactSource', 'substatus', 'groupStatus']:
    df[col] = df[col].fillna('').astype(str).str.upper().str.strip()

# Numeric
for col in ['ptpAmount', 'paymentAmount', 'OB']:
    df[col] = pd.to_numeric(df.get(col, 0), errors='coerce').fillna(0)

# ========================= FLAGS =========================

valid_substatus = (df['substatus'] != "") & (~df['substatus'].isin(EXCLUDED_SUBSTATUSES))

df['is_connected'] = (
    (df['contactSource'] == "CALL") & valid_substatus
).astype('int8')

rpc_mask = df['groupStatus'].isin(["RPC", "PTP"])

df['is_rpc'] = 0
df.loc[
    rpc_mask & ~df.duplicated(subset=['Date', 'CMS User', 'debtorId']),
    'is_rpc'
] = 1

df['is_ptp'] = (df['ptpAmount'] > 0).astype('int8')
df['is_kept'] = (df['paymentAmount'] > 0).astype('int8')

# ========================= OB =========================
df['rpc_ob']  = df['OB'] * df['is_rpc']
df['ptp_ob']  = df['OB'] * df['is_ptp']
df['kept_ob'] = df['OB'] * df['is_kept']

# ========================= AGG =========================
summary = (
    df.groupby(['Date', 'CMS User'], as_index=False)
    .agg(
        Connected_Calls=('is_connected', 'sum'),
        RPC_Count=('is_rpc', 'sum'),
        RPC_OB=('rpc_ob', 'sum'),
        PTP_Count=('is_ptp', 'sum'),
        PTP_OB=('ptp_ob', 'sum'),
        KEPT_Count=('is_kept', 'sum'),
        KEPT_OB=('kept_ob', 'sum')
    )
)

summary = summary.rename(columns={
    'Connected_Calls': 'Connected Calls',
    'RPC_Count': 'RPC Count',
    'RPC_OB': 'RPC OB',
    'PTP_Count': 'PTP Count',
    'PTP_OB': 'PTP OB',
    'KEPT_Count': 'KEPT Count',
    'KEPT_OB': 'KEPT OB'
})

summary = summary.sort_values(['Date', 'CMS User'], ascending=[False, True])

# ========================= DISPLAY =========================
st.success(f"✅ Processed {len(uploaded_files)} file(s) — {len(summary):,} summary rows")

unique_dates = summary['Date'].dropna().sort_values(ascending=False).unique()

for date in unique_dates:
    st.subheader(f"📅 {date.strftime('%B %d, %Y')}")
    daily_df = summary[summary['Date'] == date].drop(columns=['Date'])
    st.dataframe(daily_df, use_container_width=True, hide_index=True, height=450)

# ========================= EXPORT =========================
with st.sidebar:
    st.markdown("### 📥 Download Output")

    output_file = "productivity_summary_formatted.xlsx"

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        wb = writer.book
        ws = wb.create_sheet("Summary")

        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))

        row_idx = 1

        for date in unique_dates:
            # Date Header
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
            cell = ws.cell(row=row_idx, column=1, value=date.strftime('%B %d, %Y').upper())
            cell.fill = PatternFill("solid", fgColor="404040")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
            row_idx += 1

            headers = ['Date','CMS User','Connected Calls','RPC Count','RPC OB',
                       'PTP Count','PTP OB','KEPT Count','KEPT OB']

            for i, h in enumerate(headers, 1):
                c = ws.cell(row=row_idx, column=i, value=h)
                c.fill = PatternFill("solid", fgColor="808080")
                c.font = Font(color="FFFFFF", bold=True)
                c.alignment = Alignment(horizontal="center")
                c.border = thin
            row_idx += 1

            for _, r in summary[summary['Date'] == date].iterrows():
                for i, v in enumerate(r, 1):
                    c = ws.cell(row=row_idx, column=i, value=v)
                    c.border = thin
                    if isinstance(v, (int, float)) and i > 2:
                        c.number_format = '#,##0'
                row_idx += 1

            row_idx += 1

        # ✅ FIXED AUTO WIDTH (MergedCell safe)
        for col_idx in range(1, 10):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(cell.value)) for row in ws.iter_rows(min_col=col_idx, max_col=col_idx)
                 for cell in row if cell.value),
                default=0
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 25)

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    with open(output_file, "rb") as f:
        st.download_button(
            "📥 Download Formatted Excel File",
            f.read(),
            file_name=f"Productivity_Summary_{datetime.now():%Y-%m-%d_%H%M}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.caption("✅ Optimized | MergedCell error fixed | Fast & clean processing")